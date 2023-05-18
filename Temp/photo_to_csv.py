import fnmatch
import os
import re
from openpyxl import load_workbook

# открываем Excel файл
workbook = load_workbook(filename='tools.xlsx')
worksheet = workbook.active

# получаем список файлов в папке
photo_dir = 'photos'
photo_files = os.listdir(photo_dir)

# перебираем значения во второй колонке, начиная со второй строки
for cell in worksheet['B'][1:]:
    # извлекаем артикул из ячейки
    article = cell.value
    if article:
        # формируем шаблон имени файла на основе артикула
        file_pattern = f"*{article}*"
        # ищем файлы, соответствующие шаблону
        matching_files = [f for f in photo_files if fnmatch.fnmatch(f, file_pattern)]
        # проверяем, что нашлись файлы
        if matching_files:
            # записываем пути к файлам в нужные колонки
            row = cell.row
            photo1_path = os.path.join(photo_dir, matching_files[0])
            worksheet.cell(row=row, column=11, value=f"tools/{matching_files[0]}")
            if len(matching_files) > 1:
                photo2_path = os.path.join(photo_dir, matching_files[1])
                worksheet.cell(row=row, column=12, value=f"tools/{matching_files[1]}")

# сохраняем изменения в файле
workbook.save(filename='tools.xlsx')
