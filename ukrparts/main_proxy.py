import asyncio
import aiohttp
import aiofiles
from bs4 import BeautifulSoup
# from fake_useragent import UserAgent
import aiocsv
import openpyxl
import csv
from os.path import exists
# import encodings.idna

from config import Config

from aiohttp_proxy import ProxyConnector, ProxyType

config = Config()
my_proxy = f'https://{config.login}:{config.passw}@{config.ip}:{config.port}'

#user_agent = UserAgent()

#headers = {
#    'accept': '*/*',
#    'user-agent': user_agent.chrome
#}

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:107.0) Gecko/20100101 Firefox/107.0'
}

columns = ['Marka2', 'Model2', 'Pokolenie2', 'God2', 'Data Proizvodstva2', 'Obem2', 'Vid Topliva2', 'Tip div2', 'Moc2',
           'Kod Dvigat2', 'Kyzov2', 'Naimenovanie2', 'OEM2', 'Opisanie2', 'Nomer Zapchacti2', 'Proizvoditel2',
           'Risynok2', 'Url2']


def read_excel(path):
    excel_file = openpyxl.load_workbook(path)
    currently_active_sheet = excel_file.active
    res = []

    for row in currently_active_sheet.iter_rows(min_row=2, min_col=1, max_row=currently_active_sheet.max_row, max_col=2):
        row_data = [cell.value for cell in row if cell.value]
        if row_data:
            res.append(row_data)

    for title, _ in res:
        if not exists(f"{title}.csv"):
            with open(f'{title}.csv', 'w', newline='') as file:
                writer = csv.writer(file, delimiter='+')
                writer.writerow(columns)

    return res


async def scrape_product(session, url, csv_name, prod_title_data):
    async with session.get(url, headers=headers) as response:
        soup = BeautifulSoup(await response.text(), 'lxml')

        try:
            description = soup.find('ul', class_='_part_detail').find('ul').find_all('li')
            desc_res = []
            for desc in description:
                string = desc.find('span').text + ' ' + desc.text
                desc_res.append(string)

            prod_desc = '; '.join(desc_res)
            # print(f'[INFO] Description was found successfully')
        except Exception as ex:
            # print(f"[WARNING] Can't find description: {ex}/{url}")
            prod_desc = ''

        try:
            prod_photos_urls = [photo['href'] for photo in soup.find_all('a', {'rel': 'product_image'})]
            # print(f'[INFO] Photos urls were found successfully')
        except Exception as ex:
            # print(f"[WARNING] Can't find photos urls: {ex}/{url}")
            prod_photos_urls = []

        try:
            prod_original_numbers = [num.text for num in
                                     soup.find('div', string='Оригинальные (конструкторские) номера').find_next(
                                         'div', class_='row').find_all('p')]
            # print(f'[INFO] Original numbers were found successfully')
        except Exception as ex:
            # print(f"[WARNING] Can't find original numbers: {ex}/{url}")
            prod_original_numbers = []

        try:
            t = soup.find_all('table', class_='table-hover')
            if not t:
                raise AttributeError("'NoneType'")

            for table in t:
                models = table.find_all('tr', class_='showTypes')
                for model in models:
                    auto_name = model['data-manuf']
                    model_name = model['data-model']
                    data = {
                        'ID': model['data-id'],
                        'manuf': auto_name,
                        'model': model_name
                    }

                    async with session.post(
                            url='https://ukrparts.com.ua/ajax.php?act=getCompatibilityCars',
                            data=data,
                            headers=headers
                    ) as response:

                        res_soup = BeautifulSoup(await response.text(), 'lxml')

                        for tr in res_soup.find_all('tr')[1:]:
                            try:
                                prod_gen_name, prod_years = [tag.text.strip() for tag in tr.find_all('b')]
                            except:
                                prod_gen_info = [tr.text.strip() for tr in tr.find_all('td')]

                                fourwheel_drive = False

                                if '4x4 c бортовой платформой' in prod_gen_info[0]:
                                    Kyzov = '4x4 c бортовой платформой'
                                    prod_gen_info[0] = prod_gen_info[0].split()[:-4]
                                elif 'c бортовой платформой' in prod_gen_info[0]:
                                    Kyzov = 'c бортовой платформой'
                                    prod_gen_info[0] = prod_gen_info[0].split()[:-3]
                                elif '4x4' in prod_gen_info[0]:
                                    fourwheel_drive = True
                                    prod_gen_info[0] = ' '.join([i.strip() for i in prod_gen_info[0].split('4x4')])
                                    prod_gen_info[0] = prod_gen_info[0].split()
                                    Kyzov = prod_gen_info[0][-1]
                                    prod_gen_info[0] = prod_gen_info[0][:-1]
                                else:
                                    Kyzov = prod_gen_info[0].split()[-1]
                                    prod_gen_info[0] = prod_gen_info[0].split()[:-1]

                                if fourwheel_drive:
                                    Kyzov = '4x4 ' + Kyzov

                                dvig_list = []
                                Obem = ''
                                for volume in prod_gen_info[0]:
                                    try:
                                        Obem = float(volume)
                                        if 0 < Obem < 10:
                                            break
                                    except:
                                        continue

                                for item in prod_gen_info[0]:
                                    if item not in str(Obem) and item != 'i':
                                        dvig_list.append(item)

                                res_numbers = []
                                for orig in prod_original_numbers:
                                    if 'LAND ROVER' in orig or 'GREAT WALL' in orig:
                                        res_numbers.append(' '.join(orig.split()[2:]))
                                    else:
                                        res_numbers.append(' '.join(orig.split()[1:]))

                                async with aiofiles.open(f"{csv_name}.csv", mode="a+", encoding="utf-8", newline="") as afp:
                                    writer = aiocsv.AsyncWriter(afp, delimiter='+')
                                    await writer.writerow([
                                        auto_name.title(), model_name.title(), prod_gen_name,
                                        prod_gen_info[1], prod_years, Obem,
                                        prod_gen_info[-1], ' '.join(dvig_list) if type(dvig_list) is list else dvig_list,
                                        prod_gen_info[-2],
                                        prod_gen_info[2], ' '.join(Kyzov) if type(Kyzov) is list else Kyzov,
                                        prod_title_data['Naimenovanie'],
                                        ','.join(res_numbers).replace(' ', ''), prod_desc, prod_title_data['Nomer_Zapchacti'],
                                        prod_title_data['Proizvoditel'], ', '.join(prod_photos_urls), url
                                    ])
        except Exception as ex:
            # print(f"[WARNING] Can't find table of applying for auto: {ex}/{url}")

            res_numbers = []
            for orig in prod_original_numbers:
                res_numbers.append(' '.join(orig.split()[1:]))

            async with aiofiles.open(f"{csv_name}.csv", mode="a+", encoding="utf-8", newline="") as afp:
                writer = aiocsv.AsyncWriter(afp, delimiter='+')
                await writer.writerow([
                    ' ', ' ', ' ',
                    ' ', ' ', ' ',
                    ' ', ' ', ' ',
                    ' ', ' ', prod_title_data['Naimenovanie'],
                    ','.join(res_numbers).replace(' ', ''), prod_desc, prod_title_data['Nomer_Zapchacti'],
                    prod_title_data['Proizvoditel'], ', '.join(prod_photos_urls), url
                ])


async def scrape_catalogue(url, csv_name='TEST'):
    connector = ProxyConnector.from_url(my_proxy)
    async with aiohttp.ClientSession(connector=connector) as session:
        pagination_url = url
        next_page_existing = True
        pagination_counter = 1

        urls = []
        while next_page_existing:
            # print(f"PAGE {pagination_counter}")
            async with session.get(pagination_url) as response:
                soup = BeautifulSoup(await response.text(), 'lxml')

                try:
                    next_page_existing = True if 'disabled' not in soup.find('li', class_='next')['class'] else False
                    if next_page_existing:
                        pagination_url = url + soup.find('li', class_='next').find('a', class_='page-link')['href']
                        pagination_counter += 1
                except Exception as ex:
                    # print(f"[WARNING] Can't find next page button: {ex}/{url}")
                    next_page_existing = False

                try:
                    products_urls = ["https://ukrparts.com.ua" + url.find("a", class_="pointer")["href"]
                                     for url in soup.find_all('div', class_='part_box_wrapper')]
                    # print(f"[INFO] Supplies urls/titles were taken successfully")

                    prod_titles_list = []
                    for prod_title in soup.find_all('div', class_='part_box_wrapper'):
                        prod_title_data = {
                            "Naimenovanie": prod_title.find('div', class_='part_name').text.strip().split(':')[0],
                            "Proizvoditel": prod_title.find('div', class_='part_brand').text.strip(),
                            "Nomer_Zapchacti": prod_title.find('div', class_='part_article_id').text.strip()
                        }
                        prod_titles_list.append(prod_title_data)
                except Exception as ex:
                    products_urls = []
                    prod_titles_list = []
                    # print(f"[WARNING] Can't take supplies urls/titles: {ex}/{url}")

                # for product_url, prod_title_data in zip(products_urls, prod_titles_list):
                #     await scrape_product(
                #         session=session,
                #         url=product_url,
                #         csv_name=csv_name,
                #         prod_title_data=prod_title_data
                #     )

            urls.extend(zip(products_urls, prod_titles_list))

        counter = 1
        urls_count = len(urls)
        for product_url, prod_title_data in urls:
            await scrape_product(
                session=session,
                url=product_url,
                csv_name=csv_name,
                prod_title_data=prod_title_data
            )

            print(f"[{csv_name}] Обработано {counter} товаров из {urls_count}")
            counter += 1


async def main(data):
    # urls = input('Input the urls: ').split()
    tasks = []

    for csv_name, url in data:
        tasks.append(scrape_catalogue(url=url, csv_name=csv_name))

    await asyncio.gather(*tasks)


if __name__ == '__main__':
    data = read_excel('INPUT.xlsx')
    print('Данные изъяты с excel файла')
    print()
    asyncio.run(main(data))
    input('Exit')
